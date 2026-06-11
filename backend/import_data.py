"""Import 2 Excel files → normalize → seed salesbot.db (idempotent / upsert by account).

  - IR-Agreement-Status 2025-2.xlsx  (sheet 'Company visit') = master contracts
  - 20260512-PR-All-IR-List.xlsx                              = IR contacts (join on Account)

Run:  python backend/import_data.py
"""
import re
from datetime import datetime
from pathlib import Path

import openpyxl

from database import Base, SessionLocal, engine
from models import Customer

DATA_DIR = Path(__file__).resolve().parent.parent / "Data"
AGREEMENT_FILE = DATA_DIR / "IR-Agreement-Status 2025-2.xlsx"
PR_FILE = DATA_DIR / "20260512-PR-All-IR-List.xlsx"

# field -> normalized header name in the Agreement sheet
AGREEMENT_MAP = {
    "ae_ir": "ae-ir",
    "contract_type": "contract",
    "period_months": "period (month)",
    "payment_cycle": "payment",
    "monthly_payment": "current monthly payment",
    "latest_value": "latest value",
    "effective_date": "effective date",
    "expiry_date": "expiry date",
    "contract_status": "contract status",  # Active / Expired / Cancelled
    "contract_status_text": "status",       # free-text sales note e.g. "ต่อแน่นอน ไม่มีอะไร"
    "company_name_th": "company name / project name (th)",
    "company_name_en": "company name / project name (en)",
    "market": "market",
    "industry": "industry",
    "sector": "sector",
    "url": "url",
    "address_th": "address (th)",
    "address_en": "address (en)",
    "zip_code": "zip code",
}


def norm_header(h):
    """Lowercase, collapse whitespace/newlines so 'Current\\n Monthly Payment' → 'current monthly payment'."""
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h)).strip().lower()


def clean(v):
    """Normalize a cell: junk values (#VALUE!, 0-as-address, blanks) → None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.startswith("#") or s == "0":
            return None
        return s
    return v


def to_number(v):
    v = clean(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def to_int(v):
    n = to_number(v)
    return int(n) if n is not None else None


def to_dt(v):
    if isinstance(v, datetime):
        return v
    return None


def jsonify(v):
    """Make a value JSON-serializable for the `extra` column."""
    if isinstance(v, datetime):
        return v.isoformat()
    return v


def load_pr_contacts():
    """Return {account: {grade, ir_team, contact_email, cc_emails}}."""
    wb = openpyxl.load_workbook(PR_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = {norm_header(h): i for i, h in enumerate(rows[0])}
    out = {}
    for row in rows[1:]:
        acct = clean(row[header.get("account", 0)])
        if not acct:
            continue
        out[str(acct).strip().upper()] = {
            "grade": clean(row[header["grade"]]) if "grade" in header else None,
            "ir_team": clean(row[header["ir team"]]) if "ir team" in header else None,
            "contact_email": clean(row[header["email"]]) if "email" in header else None,
            "cc_emails": clean(row[header["cc."]]) if "cc." in header else None,
        }
    return out


def import_all():
    Base.metadata.create_all(bind=engine)
    contacts = load_pr_contacts()

    wb = openpyxl.load_workbook(AGREEMENT_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    raw_headers = rows[0]
    header_idx = {}
    for i, h in enumerate(raw_headers):
        nh = norm_header(h)
        if nh and nh not in header_idx:  # first occurrence wins
            header_idx[nh] = i
    acct_idx = header_idx.get("account", 1)
    mapped_indices = set(header_idx[name] for name in AGREEMENT_MAP.values() if name in header_idx)

    # An account can appear on multiple rows (contract history / renewals).
    # Collapse to one row per account = the one with the latest expiry_date (current contract).
    best = {}  # account -> (sort_key, fields, extra)
    for row in rows[1:]:
        acct_raw = clean(row[acct_idx])
        if not acct_raw:
            continue
        account = str(acct_raw).strip().upper()

        fields = {}
        for field, hname in AGREEMENT_MAP.items():
            idx = header_idx.get(hname)
            val = row[idx] if idx is not None and idx < len(row) else None
            if field in ("monthly_payment", "latest_value"):
                fields[field] = to_number(val)
            elif field == "period_months":
                fields[field] = to_int(val)
            elif field in ("effective_date", "expiry_date"):
                fields[field] = to_dt(val)
            elif field in ("address_th", "address_en"):
                fields[field] = clean(val)
            else:
                fields[field] = clean(val)

        # everything not explicitly mapped → extra (only non-null)
        extra = {}
        for nh, idx in header_idx.items():
            if idx in mapped_indices or nh == "account":
                continue
            cv = clean(row[idx]) if idx < len(row) else None
            if cv is not None:
                extra[nh] = jsonify(cv)

        # sort key: prefer latest expiry, fall back to effective date
        sort_key = fields.get("expiry_date") or fields.get("effective_date") or datetime.min
        if account not in best or sort_key >= best[account][0]:
            best[account] = (sort_key, fields, extra)

    db = SessionLocal()
    n_new = n_upd = 0
    try:
        for account, (_, fields, extra) in best.items():
            contact = contacts.get(account, {})
            existing = db.get(Customer, account)
            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                for k, v in contact.items():
                    setattr(existing, k, v)
                existing.extra = extra or None
                n_upd += 1
            else:
                db.add(Customer(account=account, extra=extra or None, **fields, **contact))
                n_new += 1
        db.commit()
    finally:
        db.close()

    print(f"Import done — unique accounts: {len(best)}, new: {n_new}, updated: {n_upd}, PR contacts: {len(contacts)}")


if __name__ == "__main__":
    import_all()
